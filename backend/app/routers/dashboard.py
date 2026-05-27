from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from app.dependencies import get_current_user
from app.models.user import UserOut
from app.db.client import get_db
import anthropic
from app.config import settings
import io
from datetime import datetime, timedelta
from typing import Any

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

# Simple in-memory cache with TTL
_cache: dict[str, tuple[Any, datetime]] = {}
CACHE_TTL_MINUTES = 15

def get_cache(key: str) -> Any | None:
  """Get cached value if not expired."""
  if key in _cache:
    value, expiry = _cache[key]
    if datetime.now() < expiry:
      return value
    del _cache[key]
  return None

def set_cache(key: str, value: Any) -> None:
  """Set cache value with TTL."""
  _cache[key] = (value, datetime.now() + timedelta(minutes=CACHE_TTL_MINUTES))

def invalidate_cache(*keys: str) -> None:
  """Invalidate cache entries."""
  for key in keys:
    _cache.pop(key, None)


async def _compute_overview():
    """Compute overview data."""
    db = get_db()
    users_count = db.table("users").select("id", count="exact").eq("is_active", True).execute()
    projects_count = db.table("projects").select("id", count="exact").eq("is_archived", False).execute()
    active_projects = (
        db.table("projects")
        .select("id", count="exact")
        .in_("status", ["approved", "in_progress"])
        .execute()
    )
    skills_count = db.table("employee_skills").select("id", count="exact").execute()

    # projects by status
    proj_all = db.table("projects").select("status").eq("is_archived", False).execute()
    status_dist: dict[str, int] = {}
    for p in (proj_all.data or []):
        s = p["status"]
        status_dist[s] = status_dist.get(s, 0) + 1

    data = {
        "total_employees": users_count.count,
        "total_projects": projects_count.count,
        "active_projects": active_projects.count,
        "total_skills_logged": skills_count.count,
        "projects_by_status": status_dist,
    }
    set_cache("overview", data)
    return data

@router.get("/overview")
async def overview(_: UserOut = Depends(get_current_user)):
    # Return cached data immediately
    cached = get_cache("overview")
    if cached:
        return cached
    # No cache, compute and return
    return await _compute_overview()


async def _compute_workforce():
    """Compute workforce data."""
    db = get_db()
    users = db.table("users").select("department, role").eq("is_active", True).limit(5000).execute()
    by_dept: dict[str, int] = {}
    by_role: dict[str, int] = {}
    for u in (users.data or []):
        dept = u.get("department") or "Unassigned"
        by_dept[dept] = by_dept.get(dept, 0) + 1
        role = u.get("role") or "employee"
        by_role[role] = by_role.get(role, 0) + 1

    # avg skill level - sample for speed
    skills = db.table("employee_skills").select("level").limit(1000).execute()
    levels = [s["level"] for s in (skills.data or [])]
    avg_level = round(sum(levels) / len(levels), 2) if levels else 0

    data = {
        "by_department": [{"name": k, "count": v} for k, v in sorted(by_dept.items(), key=lambda x: -x[1])],
        "by_role": [{"name": k, "count": v} for k, v in by_role.items()],
        "avg_skill_level": avg_level,
    }
    set_cache("workforce", data)
    return data

@router.get("/workforce")
async def workforce(_: UserOut = Depends(get_current_user)):
    cached = get_cache("workforce")
    if cached:
        return cached
    return await _compute_workforce()


async def _compute_skill_distribution():
    """Compute skill distribution data."""
    db = get_db()
    # Top skills by frequency - use simpler join
    emp_skills = db.table("employee_skills").select("skill_id, level, skills(name)").limit(5000).execute()
    skill_counts: dict[str, dict] = {}
    for es in (emp_skills.data or []):
        sid = es["skill_id"]
        skill_info = es.get("skills") or {}
        name = skill_info.get("name", sid)
        if sid not in skill_counts:
            skill_counts[sid] = {"skill_id": sid, "name": name, "domain": "General", "count": 0, "total_level": 0}
        skill_counts[sid]["count"] += 1
        skill_counts[sid]["total_level"] += es["level"]

    top_skills = sorted(skill_counts.values(), key=lambda x: -x["count"])[:15]
    for s in top_skills:
        s["avg_level"] = round(s["total_level"] / s["count"], 1) if s["count"] else 0

    # Skills by domain distribution
    by_domain: dict[str, dict] = {}
    for s in skill_counts.values():
        d = s["domain"]
        if d not in by_domain:
            by_domain[d] = {"domain": d, "count": 0, "employees": 0}
        by_domain[d]["count"] += s["count"]
        by_domain[d]["employees"] += 1

    data = {
        "top_skills": top_skills,
        "by_domain": list(by_domain.values()),
    }
    set_cache("skill_distribution", data)
    return data

@router.get("/skill-distribution")
async def skill_distribution(_: UserOut = Depends(get_current_user)):
    cached = get_cache("skill_distribution")
    if cached:
        return cached
    return await _compute_skill_distribution()


async def _compute_skill_gaps():
    """Compute skill gaps data."""
    db = get_db()
    # Active projects' required_skills vs what employees have - limit to recent projects
    active_projects = (
        db.table("projects")
        .select("id, title, rfp_extracted_data")
        .in_("status", ["approved", "in_progress", "review"])
        .eq("is_archived", False)
        .limit(50)
        .execute()
    )

    # Build demand map from RFP extracted data
    demand: dict[str, int] = {}
    for proj in (active_projects.data or []):
        rfp = proj.get("rfp_extracted_data") or {}
        required = rfp.get("required_skills") or []
        if isinstance(required, list):
            for req in required[:10]:  # limit skills per project
                skill = req.get("skill") or req.get("skill_id") or ""
                if skill:
                    demand[skill] = demand.get(skill, 0) + 1

    # Build supply map - sample
    emp_skills = db.table("employee_skills").select("skill_id, skills(name)").limit(2000).execute()
    supply: dict[str, dict] = {}
    for es in (emp_skills.data or []):
        sid = es["skill_id"]
        name = (es.get("skills") or {}).get("name", sid)
        if sid not in supply:
            supply[sid] = {"skill_id": sid, "name": name, "count": 0}
        supply[sid]["count"] += 1

    gaps = []
    for skill_name, demand_count in demand.items():
        supply_count = supply.get(skill_name, {}).get("count", 0)
        if demand_count > supply_count:
            gaps.append({
                "skill": skill_name,
                "demand": demand_count,
                "supply": supply_count,
                "gap": demand_count - supply_count,
            })

    gaps.sort(key=lambda x: -x["gap"])
    data = {"gaps": gaps[:20]}
    set_cache("skill_gaps", data)
    return data

@router.get("/skill-gaps")
async def skill_gaps(_: UserOut = Depends(get_current_user)):
    cached = get_cache("skill_gaps")
    if cached:
        return cached
    return await _compute_skill_gaps()


async def _compute_trends():
    """Compute trends data."""
    db = get_db()
    # Skill additions from audit log - limit to recent
    audit = (
        db.table("skill_audit_log")
        .select("skill_id, action, created_at, skills(name)")
        .eq("action", "add")
        .order("created_at", desc=False)
        .limit(2000)
        .execute()
    )

    month_counts: dict[str, dict[str, int]] = {}
    for entry in (audit.data or []):
        month = (entry.get("created_at") or "")[:7]
        skill_name = (entry.get("skills") or {}).get("name", entry["skill_id"])
        if month not in month_counts:
            month_counts[month] = {}
        month_counts[month][skill_name] = month_counts[month].get(skill_name, 0) + 1

    # Flatten to time series per skill
    skill_totals: dict[str, int] = {}
    for mo in month_counts.values():
        for sk, cnt in mo.items():
            skill_totals[sk] = skill_totals.get(sk, 0) + cnt

    top_trending = sorted(skill_totals.items(), key=lambda x: -x[1])[:10]

    series = []
    for skill_name, _ in top_trending:
        points = []
        for month in sorted(month_counts.keys())[-12:]:
            points.append({"month": month, "count": month_counts[month].get(skill_name, 0)})
        series.append({"skill": skill_name, "total": skill_totals[skill_name], "series": points})

    data = {"series": series}
    set_cache("trends", data)
    return data

@router.get("/trends")
async def trends(period: str = "6m", _: UserOut = Depends(get_current_user)):
    cached = get_cache("trends")
    if cached:
        return cached
    return await _compute_trends()


async def _compute_availability_overview():
    """Compute availability overview data."""
    db = get_db()
    from datetime import date
    current_month = date.today().strftime("%Y-%m")
    month_date = f"{current_month}-01"

    users = db.table("users").select("id").eq("is_active", True).limit(5000).execute()
    allocs = (
        db.table("allocations")
        .select("user_id, allocation_percentage")
        .eq("month", month_date)
        .eq("status", "confirmed")
        .limit(5000)
        .execute()
    )

    allocated: dict[str, int] = {}
    for a in (allocs.data or []):
        uid = a["user_id"]
        allocated[uid] = allocated.get(uid, 0) + a["allocation_percentage"]

    buckets = {"0-25": 0, "25-50": 0, "50-75": 0, "75-100": 0}
    for u in (users.data or []):
        avail = 100 - allocated.get(str(u["id"]), 0)
        if avail <= 25:
            buckets["0-25"] += 1
        elif avail <= 50:
            buckets["25-50"] += 1
        elif avail <= 75:
            buckets["50-75"] += 1
        else:
            buckets["75-100"] += 1

    data = {"month": current_month, "buckets": [{"range": k, "count": v} for k, v in buckets.items()]}
    set_cache("availability_overview", data)
    return data

@router.get("/availability-overview")
async def availability_overview(_: UserOut = Depends(get_current_user)):
    """Buckets employees by their current availability %."""
    cached = get_cache("availability_overview")
    if cached:
        return cached
    return await _compute_availability_overview()


@router.get("/predictions")
async def predictions(current_user: UserOut = Depends(get_current_user)):
    """AI-generated workforce skill needs forecast."""
    db = get_db()

    # Gather context - limit to recent projects and skills
    active_proj = (
        db.table("projects")
        .select("title, domain, tech_stack, rfp_extracted_data")
        .in_("status", ["approved", "in_progress", "review"])
        .limit(50)
        .execute()
    )
    top_skills_r = db.table("employee_skills").select("skill_id, skills(name)").limit(1000).execute()
    skill_freq: dict[str, int] = {}
    for es in (top_skills_r.data or []):
        name = (es.get("skills") or {}).get("name", es["skill_id"])
        skill_freq[name] = skill_freq.get(name, 0) + 1

    top_skills = sorted(skill_freq.items(), key=lambda x: -x[1])[:10]
    proj_summaries = []
    for p in (active_proj.data or [])[:10]:
        ts = ", ".join(p.get("tech_stack") or [])
        proj_summaries.append(f"- {p['title']} ({p.get('domain','')}) [{ts}]")

    prompt = (
        "You are an HR analytics expert. Based on the organization's current data:\n\n"
        f"Active projects:\n" + "\n".join(proj_summaries or ["None"]) + "\n\n"
        f"Top employee skills: {', '.join(f'{s}({c})' for s,c in top_skills)}\n\n"
        "In 3-4 bullet points, predict which skills this organization will need most in the next 6 months. "
        "Be specific and concise. Format as JSON: {\"predictions\": [\"...\", \"...\"]}"
    )

    try:
        client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        import json
        text = msg.content[0].text  # type: ignore[index]
        start = text.find("{")
        end = text.rfind("}") + 1
        data = json.loads(text[start:end])
        return data
    except Exception:
        return {"predictions": ["AI prediction unavailable — check API key configuration."]}


@router.get("/export/excel")
async def export_excel(current_user: UserOut = Depends(get_current_user)):
    """Export dashboard data as Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException, status as http_status
        raise HTTPException(status_code=http_status.HTTP_501_NOT_IMPLEMENTED, detail="openpyxl not installed")

    db = get_db()
    wb = openpyxl.Workbook()

    # Sheet 1: Employees
    ws1 = wb.active
    ws1.title = "Employees"
    headers = ["Full Name", "Email", "Role", "Department", "Job Title"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.font = Font(bold=True, color="FFFFFF")
    users = db.table("users").select("full_name, email, role, department, job_title").eq("is_active", True).execute()
    for row_idx, u in enumerate(users.data or [], 2):
        ws1.cell(row=row_idx, column=1, value=u.get("full_name"))
        ws1.cell(row=row_idx, column=2, value=u.get("email"))
        ws1.cell(row=row_idx, column=3, value=u.get("role"))
        ws1.cell(row=row_idx, column=4, value=u.get("department"))
        ws1.cell(row=row_idx, column=5, value=u.get("job_title"))

    # Sheet 2: Projects
    ws2 = wb.create_sheet("Projects")
    proj_headers = ["Title", "Status", "Domain", "Client", "Kick-off", "End Date", "Team Size"]
    for col, h in enumerate(proj_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
    projects = db.table("projects").select("title, status, domain, client_name, kick_off_date, end_date, team_size_required").eq("is_archived", False).execute()
    for row_idx, p in enumerate(projects.data or [], 2):
        ws2.cell(row=row_idx, column=1, value=p.get("title"))
        ws2.cell(row=row_idx, column=2, value=p.get("status"))
        ws2.cell(row=row_idx, column=3, value=p.get("domain"))
        ws2.cell(row=row_idx, column=4, value=p.get("client_name"))
        ws2.cell(row=row_idx, column=5, value=str(p.get("kick_off_date") or ""))
        ws2.cell(row=row_idx, column=6, value=str(p.get("end_date") or ""))
        ws2.cell(row=row_idx, column=7, value=p.get("team_size_required"))

    # Sheet 3: Skills
    ws3 = wb.create_sheet("Employee Skills")
    skill_headers = ["Employee", "Skill", "Level", "Years Experience", "Last Assessed"]
    for col, h in enumerate(skill_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
    emp_skills = db.table("employee_skills").select("level, years_experience, last_assessed_at, users(full_name), skills(name)").execute()
    for row_idx, es in enumerate(emp_skills.data or [], 2):
        ws3.cell(row=row_idx, column=1, value=(es.get("users") or {}).get("full_name"))
        ws3.cell(row=row_idx, column=2, value=(es.get("skills") or {}).get("name"))
        ws3.cell(row=row_idx, column=3, value=es.get("level"))
        ws3.cell(row=row_idx, column=4, value=es.get("years_experience"))
        ws3.cell(row=row_idx, column=5, value=str(es.get("last_assessed_at") or "")[:10])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=skillmark-export.xlsx"},
    )


@router.get("/export/pdf")
async def export_pdf(current_user: UserOut = Depends(get_current_user)):
    """Export dashboard summary as PDF with AI executive summary."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
    except ImportError:
        from fastapi import HTTPException, status as http_status
        raise HTTPException(status_code=http_status.HTTP_501_NOT_IMPLEMENTED, detail="reportlab not installed")

    db = get_db()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph("SkillMark — Dashboard Report", styles["Title"]))
    from datetime import date
    story.append(Paragraph(f"Generated: {date.today().isoformat()}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    # Overview numbers
    users_count = db.table("users").select("id", count="exact").eq("is_active", True).execute()
    projects_count = db.table("projects").select("id", count="exact").eq("is_archived", False).execute()
    active_projects = db.table("projects").select("id", count="exact").in_("status", ["approved", "in_progress"]).execute()

    story.append(Paragraph("Overview", styles["Heading2"]))
    overview_data = [
        ["Metric", "Value"],
        ["Total Employees", str(users_count.count or 0)],
        ["Total Projects", str(projects_count.count or 0)],
        ["Active Projects", str(active_projects.count or 0)],
    ]
    tbl = Table(overview_data, colWidths=[10*cm, 6*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.5*cm))

    # Top skills
    emp_skills = db.table("employee_skills").select("skill_id, skills(name)").execute()
    skill_freq: dict[str, int] = {}
    for es in (emp_skills.data or []):
        name = (es.get("skills") or {}).get("name", es["skill_id"])
        skill_freq[name] = skill_freq.get(name, 0) + 1
    top_skills = sorted(skill_freq.items(), key=lambda x: -x[1])[:10]

    story.append(Paragraph("Top Skills", styles["Heading2"]))
    skill_data = [["Skill", "Employees"]] + [[s, str(c)] for s, c in top_skills]
    skill_tbl = Table(skill_data, colWidths=[10*cm, 6*cm])
    skill_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(skill_tbl)
    story.append(Spacer(1, 0.5*cm))

    # AI Executive Summary
    story.append(Paragraph("AI Executive Summary", styles["Heading2"]))
    try:
        client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        active_proj = db.table("projects").select("title, domain, status").in_("status", ["approved", "in_progress"]).limit(10).execute()
        proj_list = ", ".join(p["title"] for p in (active_proj.data or []))
        top_skill_str = ", ".join(f"{s}({c})" for s, c in top_skills[:5])

        ai_prompt = (
            f"Write a 3-sentence executive summary for an organization dashboard report. "
            f"Active projects: {proj_list or 'none'}. Top skills: {top_skill_str or 'none'}. "
            f"Total employees: {users_count.count}. "
            "Focus on workforce strengths and recommendations."
        )
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=250,
            messages=[{"role": "user", "content": ai_prompt}],
        )
        summary_text = msg.content[0].text  # type: ignore[index]
    except Exception:
        summary_text = "AI summary unavailable."

    story.append(Paragraph(summary_text, styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=skillmark-report.pdf"},
    )
